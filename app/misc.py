import logging

import os
import sys
import time
import string

import openpyxl as xl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo



def __setup_logger(name):
    formatter = logging.Formatter(fmt='%(asctime)s %(levelname)-8s %(message)s',
                                  datefmt='%Y-%m-%d %H:%M:%S')
    handler = logging.FileHandler('log.txt', mode='w')
    handler.setFormatter(formatter)
    screen_handler = logging.StreamHandler(stream=sys.stdout)
    screen_handler.setFormatter(formatter)
    logger = logging.getLogger(name)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    logger.addHandler(screen_handler)
    return logger

logger = __setup_logger(__name__)


def safe_cast(val, to_type, default=None):
    try:
        return to_type(val)
    except (ValueError, TypeError):
        return default


def to_comma_separated_string(value):
    if isinstance(value, (list, tuple, set)):
        return ','.join(str(x) for x in value)
    return str(value)


def merge_db(db_fileprefix):
    db_wb = Workbook()
    filename = format_filename(f'{db_fileprefix}_{time.strftime("%Y%m%d-%H%M%S")}.xlsx')

    for file in os.listdir():
        if file.endswith(".xlsx") and db_fileprefix not in file:
            wb = xl.load_workbook(file)
            for ws in wb.worksheets:
                db_ws = db_wb.create_sheet(ws.title)

                max_row, max_col = ws.max_row, ws.max_column

                for i in range(1, max_row + 1):
                    for j in range(1, max_col + 1):
                        cell = ws.cell(row=i, column=j)
                        db_ws.cell(row=i, column=j).value = cell.value

                db_ws.delete_cols(1)
                db_ws.delete_rows(3)
                format_worksheet(db_ws)

    ws = db_wb.get_sheet_by_name('Sheet')
    db_wb.remove_sheet(ws)

    db_wb.save(filename)


def format_worksheet(ws):
    col, merge_to_col = 1, 1
    max_row, max_col = ws.max_row, ws.max_column

    headings = [ws.cell(row=2, column=j).value for j in range(1, max_col+1)]

    tab = Table(displayName=ws.title.replace(' ', "_"), ref=f"A2:{get_column_letter(max_col)}{max_row}")
    tab._initialise_columns()
    for column, value in zip(tab.tableColumns, headings):
        column.name = value

    while merge_to_col <= max_col:
        if ws.cell(row=1, column=merge_to_col + 1).value:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=merge_to_col)
            col = merge_to_col + 1
        merge_to_col += 1
    else:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=merge_to_col-1)

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    ws.add_table(tab)

    used_columns = (get_column_letter(column) for column in range(1, max_col+1))
    for column in used_columns:
        ws.column_dimensions[column].width = 15


def format_filename(s):
    """Take a string and return a valid filename constructed from the string.
Uses a whitelist approach: any characters not present in valid_chars are
removed. Also spaces are replaced with underscores.

Note: this method may produce invalid filenames such as ``, `.` or `..`
When I use this method I prepend a date string like '2009_01_15_19_46_32_'
and append a file extension like '.txt', so I avoid the potential of using
an invalid filename.

"""
    # Credits to @seanh from GitHub https://gist.github.com/seanh/93666
    valid_chars = "-_.() %s%s" % (string.ascii_letters, string.digits)
    filename = ''.join(c for c in s if c in valid_chars)
    filename = filename.replace(' ', '_')  # I don't like spaces in filenames.
    return filename


